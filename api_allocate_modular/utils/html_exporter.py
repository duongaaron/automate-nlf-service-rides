from datetime import datetime
import re
from zoneinfo import ZoneInfo
from bs4 import BeautifulSoup
from openpyxl import load_workbook

class HTMLExporter:
    def __init__(self):
        pass

    def clean_cell_value(self, value):
        if not isinstance(value, str):
            return value
        return re.sub(r'[^\x20-\x7E]', '', value)

    def workbook_to_html_colored(self, excel_path):
        wb = load_workbook(filename=excel_path)
        ws = wb.active

        html = '<table border="1" style="border-collapse: collapse; font-family: sans-serif;">\n'
        for row in ws.iter_rows():
            html += '<tr>'
            for cell in row:
                value = self.clean_cell_value(cell.value) if cell.value is not None else ''
                bgcolor = "FFFFFF"
                if cell.fill and cell.fill.fill_type == "solid":
                    raw_color = cell.fill.start_color.rgb
                    if raw_color:
                        bgcolor = raw_color[-6:]  # Strip alpha
                is_bold = "font-weight: bold;" if isinstance(value, str) and value.strip().startswith("Driver:") else ""
                html += f'<td style="background-color: #{bgcolor}; padding: 5px; {is_bold}">{value}</td>'
            html += '</tr>\n'
        html += '</table>'
        return html
    
    def export(self, data):
        excel_path = data.get("excel_path")
        if not excel_path:
            print("[HTMLExporter] No Excel path provided.")
            return

        assignments_html_content = self.workbook_to_html_colored(excel_path)

        # Extract just the filename from the path
        excel_filename = excel_path.split("/")[-1] if "/" in excel_path else excel_path

        with open("./api_allocate_modular/outputs/html/assignments_table.html", "w", encoding="utf-8") as f:
            f.write(assignments_html_content)

        now_cst = datetime.now(ZoneInfo("America/Chicago"))
        formatted_time = now_cst.strftime("%m-%d-%Y %H-%M-%S")
            
        excel_filename = f"assignments_{formatted_time}.xlsx"

        rides_to_html = "../maps/rides_to"
        rides_back_html = "../maps/rides_back"

        index_html = f"""
        <h1>Click on the links to show the map of partitions for this week's driver/rider assignments. Updated {formatted_time} CST</h1>
        <ul>
        <li><a href="{rides_to_html}">Rides to map</a></li>
        <li><a href="{rides_back_html}">Rides back map</a></li>
        </ul>

        <div style="z-index: 9999; display: flex; gap: 10px;">
        <a href="{excel_filename}" download>
            <button style="margin: 10px; padding: 10px 20px; font-size: 14px; background-color: #4CAF50; color: white; border: none; border-radius: 5px;">
            Download Assignments as Excel
            </button>
        </a>
        </div>

        <div style="overflow-x: auto; -webkit-overflow-scrolling: touch; touch-action: auto;">
        <iframe
            src="assignments_table.html"
            width="100%"
            style="min-width: 1000px; height: 80vh; border: 1px solid #ccc;"
        ></iframe>
        </div>
        """

        # Load the HTML file
        with open("./api_allocate_modular/outputs/html/index.html", "w", encoding="utf-8") as f:
            f.write(index_html)

        print("[HTMLExporter] HTML table saved to assignments_table.html")
        data["html_path"] = "assignments_table.html"