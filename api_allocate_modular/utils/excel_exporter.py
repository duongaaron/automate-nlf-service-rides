import os
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from utils.constants import location_colors, EVENT_TYPES

# ✅ Moved outside the class
def address_key(address):
    return address if address in location_colors else address.split()[0]

class ExcelExporter:
    def __init__(self, output_dir="."):
        self.output_dir = output_dir

    def export(self, data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Assignments"
        formatted_time = data.get("formatted_time", datetime.now().strftime("%Y-%m-%d_%H-%M"))

        used_cols = set()
        start_col_to = 3
        start_col_back = 10
        key_col = 1
        row_offset = 0

        for event_key in EVENT_TYPES:
            label = EVENT_TYPES[event_key]["label"]

            # Section header
            ws.merge_cells(start_row=1 + row_offset, start_column=start_col_to, end_row=1 + row_offset, end_column=start_col_back + 4)
            header_cell = ws.cell(row=1 + row_offset, column=start_col_to, value=f"{label} Assignments")
            header_cell.font = Font(bold=True, size=14)

            # TO rides
            used_cols |= self._place_assignments(
                ws,
                data.get(f"assignments_to_{event_key}", {}),
                data.get(f"unassigned_riders_to_{event_key}", []),
                data.get("oc_people_w_invalid_address", []),
                start_col=start_col_to,
                key_col=key_col,
                sort_key_driver=lambda d: (d.service_type, d.pickup_location.split()[0]),
                sort_key_rider=lambda r: r.pickup_location.split()[0],
                driver_color_key=lambda d: address_key(d.pickup_location),
                rider_color_key=lambda r: address_key(r.pickup_location),
                label_plan=False,
                include_rider_keys_in_legend=True,
                row_offset=row_offset + 1,
                section_label="To"
            )

            # BACK rides
            used_cols |= self._place_assignments(
                ws,
                data.get(f"assignments_back_{event_key}", {}),
                data.get(f"unassigned_riders_back_{event_key}", []),
                data.get("oc_people_w_invalid_address", []),
                start_col=start_col_back,
                key_col=key_col,
                sort_key_driver=lambda d: d.plans,
                sort_key_rider=lambda r: r.pickup_location.split()[0],
                driver_color_key=lambda d: address_key(d.plans),
                rider_color_key=lambda r: address_key(r.pickup_location),
                label_plan=True,
                include_rider_keys_in_legend=False,
                row_offset=row_offset + 1,
                section_label="Back"
            )

            row_offset += 30  # space between events

        # Auto-size columns
        for col_index in used_cols.union({key_col}):
            letter = get_column_letter(col_index)
            max_length = max((len(str(cell.value)) for cell in ws[letter] if cell.value), default=0)
            ws.column_dimensions[letter].width = max(15, max_length + 2)

        filename = f"assignments_{formatted_time}.xlsx"
        local_path = os.path.join(self.output_dir, filename)

        # 🔥 Delete all existing assignments*.xlsx files in the output directory
        for f in glob.glob(os.path.join(self.output_dir, "assignments*.xlsx")):
            try:
                os.remove(f)
            except Exception as e:
                print(f"Warning: Could not delete file {f}: {e}")
        wb.save(local_path)
        print(f"[ExcelExporter] Saved to {local_path}")

        for event_key in EVENT_TYPES:
            for direction in ["to", "back"]:
                subfolder = f"./maps/rides_{direction}_{event_key}"
                os.makedirs(subfolder, exist_ok=True)
                full_copy_path = os.path.join(subfolder, filename)
                for f in glob.glob(os.path.join(subfolder, "assignments*.xlsx")):
                    try:
                        os.remove(f)
                    except:
                        pass
                wb.save(full_copy_path)
                print(f"[ExcelExporter] Copied to {full_copy_path}")

        data["excel_path"] = local_path
        return local_path

    def _place_assignments(
        self, ws, assignments, unassigned_riders, oc_people_w_invalid_address,
        start_col, key_col, sort_key_driver, sort_key_rider,
        driver_color_key, rider_color_key,
        label_plan, include_rider_keys_in_legend,
        row_offset, section_label
    ):
        row = row_offset + 1
        local_used_cols = set()
        used_keys = set()
        DRIVERS_PER_ROW = 4

        sorted_assignments = sorted(assignments.items(), key=lambda item: sort_key_driver(item[0]))
        driver_groups = [sorted_assignments[i:i + DRIVERS_PER_ROW] for i in range(0, len(sorted_assignments), DRIVERS_PER_ROW)]

        for group in driver_groups:
            max_height = 0
            col = start_col

            for driver, riders in group:
                driver_text = f"{section_label}: {driver.name}"
                if label_plan:
                    driver_text += f" [{driver.plans}]"
                driver_cell = ws.cell(row=row, column=col, value=driver_text)
                driver_key = driver_color_key(driver)
                used_keys.add(driver_key)
                fill_color = location_colors.get(driver_key, "FFFFFF")
                driver_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                driver_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                driver_cell.font = Font(bold=True, underline="single")
                local_used_cols.add(col)

                sorted_riders = sorted(riders, key=sort_key_rider)
                for i, rider in enumerate(sorted_riders):
                    rider_text = f"{rider.name}"
                    rider_cell = ws.cell(row=row + 1 + i, column=col, value=rider_text)
                    rider_key = rider_color_key(rider)
                    if include_rider_keys_in_legend:
                        used_keys.add(rider_key)
                    fill_color = location_colors.get(rider_key, "FFFFFF")
                    rider_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    local_used_cols.add(col)

                max_height = max(max_height, len(sorted_riders))
                col += 1

            row += max_height + 3

        if unassigned_riders or oc_people_w_invalid_address:
            col += 1
            header_cell = ws.cell(row=row, column=col, value="UNASSIGNED RIDERS")
            header_cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center")
            header_cell.font = Font(bold=True, underline="single", color="FF0000")
            local_used_cols.add(col)

            all_unassigned = sorted(set(unassigned_riders).union(oc_people_w_invalid_address), key=sort_key_rider)
            for i, rider in enumerate(all_unassigned):
                base = f"{rider.name}"
                if rider in unassigned_riders:
                    base += " — No valid driver"
                else:
                    base += " — Invalid off-campus address"
                cell = ws.cell(row=row + 1 + i, column=col, value=base)
                rider_key = rider_color_key(rider).strip()
                if include_rider_keys_in_legend:
                    used_keys.add(rider_key)
                fill_color = location_colors.get(rider_key, "FFFFFF")
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                local_used_cols.add(col)

        # Legend
        key_row = 2
        for key in sorted(used_keys):
            key_cell = ws.cell(row=key_row, column=key_col, value=key)
            fill_color = location_colors.get(key, "FFFFFF")
            key_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            key_row += 1

        return local_used_cols